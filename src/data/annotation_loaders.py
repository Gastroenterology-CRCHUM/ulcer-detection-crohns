"""Annotation Excel loaders for each colonoscopy pipeline."""

from __future__ import annotations

import logging
import random
import re
from pathlib import Path

import pandas as pd

logger = logging.getLogger(__name__)


def _hms_to_seconds(hms) -> float | None:
    """Convert HH:MM:SS string or timedelta to total seconds."""
    if pd.isna(hms):
        return None
    if hasattr(hms, "total_seconds"):
        return float(hms.total_seconds())
    try:
        parts = str(hms).strip().split(":")
        h, m, s = int(parts[0]), int(parts[1]), float(parts[2])
        return h * 3600 + m * 60 + s
    except (ValueError, IndexError):
        return None


def parse_timestamp(ts: str) -> float:
    """Parse a timestamp string to seconds.

    Accepts HH:MM:SS[.fff] and MM:SS[.fff] formats.

    Raises:
        ValueError: If the string cannot be parsed.
    """
    ts = str(ts).strip().replace(",", ".")
    m = re.fullmatch(r"(\d+):(\d{2}):(\d{2}(?:\.\d+)?)", ts)
    if m:
        return int(m.group(1)) * 3600 + int(m.group(2)) * 60 + float(m.group(3))
    m = re.fullmatch(r"(\d+):(\d{2}(?:\.\d+)?)", ts)
    if m:
        return int(m.group(1)) * 60 + float(m.group(2))
    raise ValueError(f"Cannot parse timestamp: {ts!r}")


def sample_intervals_from_excel(
    excel_path: Path,
    record_id: str,
    fps: float,
    interval_sec: float,
    seed: int | None = None,
) -> list[tuple[int, int, str]]:
    """Sample one random sub-interval per annotation segment from an Excel file.

    Reads "Informative Timestamps" and "Non-Informative Timestamps" sheets.
    For each segment belonging to *record_id* whose duration > *interval_sec*,
    picks one random sub-interval of *interval_sec* seconds inside it.

    Args:
        excel_path:   Path to annotation Excel workbook.
        record_id:    Video record ID to filter rows.
        fps:          Video frame rate (used to convert seconds → frames).
        interval_sec: Length of each sampled sub-interval in seconds.
        seed:         Random seed for reproducibility.

    Returns:
        List of ``(start_frame, end_frame, category_hint)`` tuples.
        *category_hint* is ``"Informative"`` or ``"Non-Informative"``.

    Raises:
        ImportError: If openpyxl is not installed.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("pip install openpyxl") from None

    rng = random.Random(seed)
    results: list[tuple[int, int, str]] = []

    wb = openpyxl.load_workbook(excel_path, data_only=True)

    for sheet_name in ("Informative Timestamps", "Non-Informative Timestamps"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header = [
            str(c.value).strip().lower() if c.value else ""
            for c in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            d = {header[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)}
            if d.get("record_id", "").lower() != record_id.lower():
                continue
            try:
                t_start = parse_timestamp(d["start_time"])
                t_end = parse_timestamp(d["end_time"])
            except (KeyError, ValueError):
                continue

            duration = t_end - t_start
            if duration <= interval_sec:
                continue

            max_offset = duration - interval_sec
            offset = rng.uniform(0, max_offset)
            sub_start = t_start + offset
            sub_end = sub_start + interval_sec

            results.append(
                (
                    int(sub_start * fps),
                    int(sub_end * fps),
                    sheet_name.split()[0],  # "Informative" or "Non-Informative"
                )
            )

    return results


def load_ulcer_annotations(excel_path: Path) -> pd.DataFrame:
    """Load ulcer/non-ulcer timestamps from Excel → unified DataFrame.

    Expected sheets: "Ulcer timestamps", "Non-Ulcer timestamps".
    Expected columns: record_id, start_time (HH:MM:SS), end_time, sample_number, Size:.

    Returns
    -------
    DataFrame with columns: record_id, start_s, end_s, sample_number, size, label
        label: 1 = ulcer, 0 = non-ulcer
    """
    dfs = []
    for sheet, label in (("Ulcer timestamps", 1), ("Non-Ulcer timestamps", 0)):
        try:
            df = pd.read_excel(excel_path, sheet_name=sheet)
        except Exception as exc:
            raise ValueError(f"Cannot read sheet '{sheet}' from {excel_path}: {exc}") from exc
        df = df.rename(columns={"start_time": "start_hms", "end_time": "end_hms", "Size:": "size"})
        df["label"] = label
        dfs.append(df)

    out = pd.concat(dfs, ignore_index=True)
    out["start_s"] = out["start_hms"].apply(_hms_to_seconds)
    out["end_s"] = out["end_hms"].apply(_hms_to_seconds)
    out = out.dropna(subset=["record_id", "start_s", "end_s"]).copy()
    out = out[out["end_s"] > out["start_s"]].copy()
    out["record_id"] = out["record_id"].astype(str).str.strip()
    out = out.sort_values(["record_id", "start_s"]).reset_index(drop=True)
    present = [
        c
        for c in ["record_id", "start_s", "end_s", "sample_number", "size", "label"]
        if c in out.columns
    ]
    return out[present]
